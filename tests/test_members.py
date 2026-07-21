from io import BytesIO

from openpyxl import load_workbook


def _seed_members(app):
    from app.extensions import db
    from app.models import MemberProfile, User

    with app.app_context():
        manager = User(
            auth_user_id=1,
            username='manager-user',
            display_name='Manager User',
            email='manager@example.com',
            platform_role='admin',
            service_role='user',
            profile_complete=True,
        )
        member = User(
            auth_user_id=77,
            username='maxmuster',
            display_name='Max Muster',
            email='max@example.com',
            platform_role='user',
            service_role='user',
            profile_complete=True,
        )
        member_two = User(
            auth_user_id=78,
            username='zoespieler',
            display_name='Zoe Spieler',
            email='zoe@example.com',
            platform_role='user',
            service_role='user',
            profile_complete=False,
        )
        db.session.add(manager)
        db.session.add(member)
        db.session.add(member_two)
        db.session.flush()
        db.session.add(MemberProfile(
            user_id=member.id,
            first_name='Max',
            last_name='Muster',
            position='OL',
            jersey_number='12',
            shirt_size='L',
            phone='+41 79 111 11 11',
            address_line1='Musterstrasse 1',
            postal_code='3600',
            city='Thun',
            nationality='Schweiz',
        ))
        db.session.add(MemberProfile(user_id=member_two.id, first_name='Zoe', last_name='Spieler', position='DB', jersey_number='3', shirt_size='M'))
        db.session.commit()
        return manager.id, member.id, member_two.id


def _fake_members_payload():
    return {
        'teams': [],
        'is_platform_admin': True,
        'users': [
            {
                'id': 77,
                'username': 'maxmuster',
                'display_name': 'Max Muster',
                'email': 'max@example.com',
                'account_status': 'active',
                'profile_complete': True,
                'active_memberships': [
                    {
                        'team': {'name': 'U18', 'code': 'U18', 'id': 10},
                        'member_role': 'player',
                    },
                ],
                'pending_memberships': [
                    {
                        'team': {'name': 'Senioren', 'code': 'SR', 'id': 11},
                        'member_role': 'coach',
                    },
                ],
            },
            {
                'id': 78,
                'username': 'zoespieler',
                'display_name': 'Zoe Spieler',
                'email': 'zoe@example.com',
                'account_status': 'active',
                'profile_complete': False,
                'active_memberships': [
                    {
                        'team': {'name': 'U16', 'code': 'U16', 'id': 12},
                        'member_role': 'player',
                    },
                ],
                'pending_memberships': [],
            },
        ],
    }


def test_members_page_renders_compact_status_badges(client, app, monkeypatch):
    from app.routes import main as main_routes

    manager_id, _, _ = _seed_members(app)
    monkeypatch.setattr(main_routes, '_fetch_members_for_manager', lambda user: (_fake_members_payload(), None))

    with client.session_transaction() as session:
        session['user_id'] = manager_id

    response = client.get('/members')
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert 'Mannschaftsliste als XLSX' in html
    assert 'bi-patch-check-fill' in html
    assert 'bi-pencil' in html
    assert 'Max' in html
    assert 'Muster' in html
    assert '+41 79 111 11 11' in html
    assert 'max@example.com' in html
    assert 'Musterstrasse 1, 3600 Thun' in html
    assert 'Profil vollständig' not in html
    assert 'Jersey' not in html


def test_members_export_returns_xlsx(client, app, monkeypatch):
    from app.routes import main as main_routes

    manager_id, _, _ = _seed_members(app)
    monkeypatch.setattr(main_routes, '_fetch_members_for_manager', lambda user: (_fake_members_payload(), None))

    with client.session_transaction() as session:
        session['user_id'] = manager_id

    response = client.get('/members/export')
    workbook = load_workbook(filename=BytesIO(response.data))
    sheet = workbook.active

    assert response.status_code == 200
    assert response.headers['Content-Disposition'].endswith('mannschaftsliste.xlsx')
    assert response.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    assert sheet.title == 'Mannschaftsliste'
    assert [cell.value for cell in sheet[1]] == [
        'Vorname',
        'Nachname',
        'Position',
        'Aktive Teams',
        'Aktive Rollen',
        'Status',
        'E-Mail',
        'Telefon',
        'Adresse 1',
        'Adresse 2',
        'PLZ',
        'Ort',
        'Nationalität',
        'Geburtsdatum',
        'Offene Anfragen',
    ]
    rows = [
        [sheet.cell(row=row, column=index).value for index in range(1, 16)]
        for row in range(2, sheet.max_row + 1)
    ]
    assert len(rows) == 2
    assert rows[0] == [
        'Max',
        'Muster',
        'OL',
        'U18',
        'Spieler',
        'ACTIVE',
        'max@example.com',
        '+41 79 111 11 11',
        'Musterstrasse 1',
        None,
        '3600',
        'Thun',
        'Schweiz',
        None,
        'Senioren: Coach',
    ]
    assert rows[1] == [
        'Zoe',
        'Spieler',
        'DB',
        'U16',
        'Spieler',
        'ACTIVE',
        'zoe@example.com',
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    ]


def test_members_export_returns_team_filtered_xlsx(client, app, monkeypatch):
    from app.routes import main as main_routes

    manager_id, _, _ = _seed_members(app)
    monkeypatch.setattr(main_routes, '_fetch_members_for_manager', lambda user: (_fake_members_payload(), None))

    with client.session_transaction() as session:
        session['user_id'] = manager_id

    response = client.get('/members/export?team=U18')
    workbook = load_workbook(filename=BytesIO(response.data))
    sheet = workbook.active

    assert response.status_code == 200
    assert response.headers['Content-Disposition'].endswith('mannschaftsliste_U18.xlsx')
    assert sheet.max_row == 2
    assert sheet.cell(row=2, column=1).value == 'Max'